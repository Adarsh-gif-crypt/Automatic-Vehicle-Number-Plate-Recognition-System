# Currently static Can be scheduled using schedule python package 

import os
from openpyxl import Workbook,load_workbook
from openpyxl.styles import numbers
import numpy as np
from matplotlib import pyplot as plt
import cv2 as cv
import pytesseract
import streamlit as st
from PIL import Image
pytesseract.pytesseract.tesseract_cmd = r"C:\Program Files\Tesseract-OCR\tesseract.exe"

st.set_page_config(layout='wide')

titles0, title, titles1 = st.columns((1.1,6,0.5))
title.title('Automatic Number Plate Recognition System')

image1, image2 = st.columns(2)
image2.image('https://trinitycctv.co.nz/wp-content/uploads/2016/11/Website-gallery-6.jpg')
image1.image('https://1.bp.blogspot.com/-y7pZG5gYNQk/YTH8EpXIdDI/AAAAAAAAEr0/QlEsiI6qOdQopAsaZUOTkTPxaCQvIIbMgCLcBGAsYHQ/s1200/MP%2BPolice%2BRecruitment-min.jpg')


rows0,rowc1,rowc2 = st.columns(3)
st.markdown("""##### A system that detects License Plates from moving vehicles. To get started , upload your image file's path in the box given below""")
rowc1.caption('Only intended for use by Madhya Pradesh Police Department®')

# uploaded_file = st.file_uploader('Choose a file')
# if uploaded_file is not None:
#     path_in = uploaded_file.name
#     st.write(path_in)
# else:
#     path_in = None

filepath1 = st.text_input('Enter File Path',placeholder='File Path')
# st.write(filepath)




#if(category == 'Test 1')



def process_image(cropped_plate):
    cropped_image = cv.imread(cropped_plate)
    cropped_image = cv.cvtColor(cropped_image, cv.COLOR_BGR2RGB)
    pixel_vals = cropped_image.reshape((-1,3))
    pixel_vals = np.float32(pixel_vals)
    criteria = (cv.TERM_CRITERIA_EPS+cv.TERM_CRITERIA_MAX_ITER, 100, 0.85)
    k = 2
    retval, labels, centers = cv.kmeans(pixel_vals, k, None, criteria, 10, cv.KMEANS_RANDOM_CENTERS)
    centers = np.uint8(centers)
    meta = centers[labels.flatten()]
    cropped_image = meta.reshape((cropped_image.shape))
    kernel = np.ones((5,5),np.float32)/25
    dst = cv.filter2D(cropped_image,-1,kernel)

    text_from_img = pytesseract.image_to_string(dst)
    return text_from_img

#raw_img = r'{}'.format(filepath1)
#image = Image.open(raw_img)
#st.image(image, caption = 'Uploaded Image')


if st.button('Detect and Save'):
    raw_s = r'{}'.format(filepath1)
    filepath=raw_s
    os.chdir(r'./yolov5')
    status=os.system(r'python detect.py --source {file} --weights runs/train/exp2/weights/best.pt'.format(file=filepath))       # model 
    status=0 #--------------------
    if(status==0):
        os.chdir(r'..')
        if 1:
            #no workbook
            #create
            pass
        workbook = load_workbook(filename="./final.xlsx")
        sheet = workbook.active
        max_rows=sheet.max_row
        row=1


        # '''image processing, text extraction
        # move to excel
        # delete files / move to database'''


        for filename in os.listdir(r'./clean_plates'):
            plate_number= process_image(os.path.join(r'./clean_plates',filename)) #process_image(filepath)        #process image , get text
            filename=filename[:-4]
            date=filename[:-4].split('--')[0]          # Extract date and time
            time=filename[:-4].split('--')[1]
            time=time.replace('-',':')
            date_time=date+' '+time
            if len(plate_number)==0:
                continue

            item=[]
            item.append(date_time)
            item.append(plate_number)

            sheet.append(item)                                                                  # Add data to workbook
            sheet.cell(row=row+max_rows,column=1).number_format=numbers.FORMAT_DATE_DATETIME   # change number formatting
            row+=1

            workbook.save(filename="final.xlsx")       # save the file
            os.remove(os.path.join('./clean_plates',filename+'.jpg'))   # remove pictures, can also be moved to a database              

            
            stream = os.popen('node put-files.js --token=eyJhbGciOiJIUzI1NiIsInR5cCI6IkpXVCJ9.eyJzdWIiOiJkaWQ6ZXRocjoweDczMTlFNWY1RjlBQ2IwZUU3MjQyRjExNjNEMjlDNTFmMWI0MjQzQjAiLCJpc3MiOiJ3ZWIzLXN0b3JhZ2UiLCJpYXQiOjE2NDc1MjQ4NzA3MjEsIm5hbWUiOiJhbnByIn0.ETARUpu2rIp--p5nIrUnpWRNmOl2un7_tH7vorRqao8 final.xlsx')
            output = stream.read()
            if output is not None:
                st.write(output)
                st.write('Copy this CID for access to the Secured file')
#API_token = 'eyJhbGciOiJIUzI1NiIsInR5cCI6IkpXVCJ9.eyJzdWIiOiJkaWQ6ZXRocjoweDczMTlFNWY1RjlBQ2IwZUU3MjQyRjExNjNEMjlDNTFmMWI0MjQzQjAiLCJpc3MiOiJ3ZWIzLXN0b3JhZ2UiLCJpYXQiOjE2NDc1MjQ4NzA3MjEsIm5hbWUiOiJhbnByIn0.ETARUpu2rIp--p5nIrUnpWRNmOl2un7_tH7vorRqao8'


# process = subprocess.Popen(['node put-files.js --token=eyJhbGciOiJIUzI1NiIsInR5cCI6IkpXVCJ9.eyJzdWIiOiJkaWQ6ZXRocjoweDczMTlFNWY1RjlBQ2IwZUU3MjQyRjExNjNEMjlDNTFmMWI0MjQzQjAiLCJpc3MiOiJ3ZWIzLXN0b3JhZ2UiLCJpYXQiOjE2NDc1MjQ4NzA3MjEsIm5hbWUiOiJhbnByIn0.ETARUpu2rIp--p5nIrUnpWRNmOl2un7_tH7vorRqao8 final.xlsx'],shell = True, stdout = subprocess.PIPE)
# stdout, stderr = process.communicate()
# print(process, stderr)

st.caption('Made by Team Greninja')
#### END OF PROGRAM ####